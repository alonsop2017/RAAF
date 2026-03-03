#!/usr/bin/env python3
"""
gmail_invoices.py — Fetch Anthropic API invoice emails from Gmail.

Searches the last 90 days for Anthropic billing/invoice emails,
skips Claude.ai subscription invoices, and summarizes API usage invoices.

FIRST-TIME SETUP (one-time only):
  1. In Google Cloud Console → APIs & Services → Credentials,
     open your OAuth 2.0 client and add this to Authorized redirect URIs:
       http://localhost:8765
  2. If running on a remote server (headless), open a NEW local terminal and run:
       ssh -L 8765:localhost:8765 alonsop@raaf.genapex.org
     Leave that tunnel open, then run this script.

USAGE:
  python scripts/gmail_invoices.py

Subsequent runs use the saved token (config/.gmail_token.json) — no browser needed.
"""

import sys
import os
import json
import base64
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.stdout.reconfigure(encoding="utf-8")

import yaml
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# ── Config ───────────────────────────────────────────────────────────────────

SCOPES        = ["https://www.googleapis.com/auth/gmail.readonly"]
TOKEN_PATH    = PROJECT_ROOT / "config" / ".gmail_token.json"
SETTINGS_PATH = PROJECT_ROOT / "config" / "settings.yaml"
OAUTH_PORT    = 8765
LOOKBACK_DAYS = 90

# Email body / subject patterns that identify Claude.ai subscription invoices
# (not API invoices — we skip these)
SUBSCRIPTION_KEYWORDS = [
    "claude.ai",
    "claude pro",
    "claude max",
    "claude subscription",
    "monthly subscription",
    "consumer subscription",
    "claude.ai subscription",
]

# ── Credentials ──────────────────────────────────────────────────────────────

def _load_client_config() -> dict:
    """Read client_id and client_secret from settings.yaml + env."""
    with open(SETTINGS_PATH) as f:
        settings = yaml.safe_load(f)
    client_id  = settings["auth"]["google"]["client_id"]
    secret_env = settings["auth"]["google"].get("client_secret_env", "GOOGLE_CLIENT_SECRET")
    client_secret = os.environ.get(secret_env, "")
    if not client_secret:
        print(f"\nERROR: environment variable '{secret_env}' is not set.")
        print("Export it before running:  export GOOGLE_CLIENT_SECRET=<your-secret>")
        sys.exit(1)
    return {
        "installed": {
            "client_id": client_id,
            "client_secret": client_secret,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "redirect_uris": [f"http://localhost:{OAUTH_PORT}"],
        }
    }


def get_credentials() -> Credentials:
    """Return valid Gmail credentials, refreshing or re-authorizing as needed."""
    creds = None

    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if creds and creds.valid:
        return creds

    if creds and creds.expired and creds.refresh_token:
        print("Refreshing access token...")
        creds.refresh(Request())
        TOKEN_PATH.write_text(creds.to_json())
        return creds

    # ── First-time authorization ──────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  Gmail Authorization Required")
    print("=" * 60)
    print(f"""
BEFORE continuing, make sure:

  1. http://localhost:{OAUTH_PORT} is added as an Authorized Redirect URI
     in Google Cloud Console for your OAuth client.

  2. If this server is headless, open a new LOCAL terminal and run:
       ssh -L {OAUTH_PORT}:localhost:{OAUTH_PORT} alonsop@raaf.genapex.org
     Leave that SSH tunnel open.

Press Enter when ready...""")
    input()

    flow = InstalledAppFlow.from_client_config(_load_client_config(), SCOPES)
    creds = flow.run_local_server(
        port=OAUTH_PORT,
        open_browser=True,
        success_message="Authorization complete. You can close this tab.",
    )
    TOKEN_PATH.write_text(creds.to_json())
    print(f"\nToken saved → {TOKEN_PATH.relative_to(PROJECT_ROOT)}")
    return creds

# ── Email parsing ─────────────────────────────────────────────────────────────

def _decode_body(payload: dict) -> str:
    """Recursively extract readable text from a Gmail message payload."""
    mime = payload.get("mimeType", "")
    data = payload.get("body", {}).get("data", "")

    if mime == "text/plain" and data:
        return base64.urlsafe_b64decode(data).decode("utf-8", errors="replace")

    if mime == "text/html" and data:
        html = base64.urlsafe_b64decode(data).decode("utf-8", errors="replace")
        # Strip tags and collapse whitespace for a rough plain-text view
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"[ \t]{2,}", " ", text)
        return text

    for part in payload.get("parts", []):
        result = _decode_body(part)
        if result.strip():
            return result

    return ""


def _is_subscription(subject: str, body: str) -> bool:
    combined = (subject + " " + body[:2000]).lower()
    return any(kw in combined for kw in SUBSCRIPTION_KEYWORDS)


def _find_amount(body: str) -> str:
    patterns = [
        r"Total[^\n$]*\$[\d,]+\.\d{2}",
        r"Amount[^\n$]*\$[\d,]+\.\d{2}",
        r"\$[\d,]+\.\d{2}",
        r"USD\s*[\d,]+\.\d{2}",
    ]
    for pat in patterns:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return "—"


def _find_invoice_id(subject: str, body: str) -> str:
    patterns = [
        r"Invoice[:\s#]+([A-Z0-9\-_]+)",
        r"Receipt[:\s#]+([A-Z0-9\-_]+)",
        r"#\s*([A-Z0-9\-]{6,})",
    ]
    for pat in patterns:
        m = re.search(pat, subject + " " + body[:1000], re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return "—"


def _find_period(body: str) -> str:
    """Try to extract billing period from email body."""
    patterns = [
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2}[–\-,\s]+\w+\s+\d{1,2},?\s*\d{4}",
        r"billing period[:\s]+([^\n]{5,40})",
        r"period[:\s]+([^\n]{5,40})",
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}",
    ]
    for pat in patterns:
        m = re.search(pat, body, re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return "—"


def _clean_preview(body: str, max_chars: int = 800) -> list[str]:
    """Return meaningful non-empty lines from the body for preview."""
    lines = []
    for line in body[:max_chars].splitlines():
        line = line.strip()
        if len(line) > 3:
            lines.append(line)
    return lines

# ── Excel export ──────────────────────────────────────────────────────────────

def _parse_amount_float(amount_str: str) -> float | None:
    """Extract a numeric float from an amount string like 'Total $5.38' or 'CA$5.00'."""
    m = re.search(r"[\d,]+\.\d{2}", amount_str)
    if m:
        return float(m.group(0).replace(",", ""))
    return None


def _parse_date(date_str: str) -> datetime | None:
    """Parse RFC 2822 email date to datetime."""
    from email.utils import parsedate_to_datetime
    try:
        return parsedate_to_datetime(date_str)
    except Exception:
        return None


def export_excel(api_invoices: list[dict], out_path: Path) -> None:
    """Write invoice list to a formatted Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anthropic API Invoices"

    # ── Styles ────────────────────────────────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1F3864")
    total_fill     = PatternFill("solid", fgColor="D9E1F2")
    total_font     = Font(bold=True, size=11)
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="CCCCCC")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_fmt   = '#,##0.00'

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = ["#", "Date", "Receipt #", "Invoice #", "Description", "Currency", "Amount (numeric)", "Amount (raw)"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, inv in enumerate(api_invoices, 1):
        dt       = _parse_date(inv["date"])
        date_str = dt.strftime("%Y-%m-%d %H:%M UTC") if dt else inv["date"]

        # Receipt # comes from the subject line: "... #XXXX-XXXX-XXXX"
        receipt_m = re.search(r"#([\d\-]+)", inv["subject"])
        receipt   = receipt_m.group(1) if receipt_m else "—"

        # Currency: CA$ vs $
        currency  = "CAD" if "CA$" in inv["amount"] or "CA$" in inv.get("preview", [""])[0] else "USD"
        amount_f  = _parse_amount_float(inv["amount"])

        row = [
            i,
            date_str,
            receipt,
            inv["invoice_id"] if inv["invoice_id"] != "illustration" else "—",
            inv["subject"],
            currency,
            amount_f,
            inv["amount"],
        ]
        ws.append(row)
        r = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border    = border
            cell.alignment = center if col in (1, 6) else left
            if col == 7 and amount_f is not None:
                cell.number_format = currency_fmt

    # ── Totals row ────────────────────────────────────────────────────────────
    usd_total = sum(
        _parse_amount_float(inv["amount"]) or 0
        for inv in api_invoices
        if "CA$" not in inv["amount"]
    )
    cad_total = sum(
        _parse_amount_float(inv["amount"]) or 0
        for inv in api_invoices
        if "CA$" in inv["amount"]
    )

    last_data = ws.max_row
    ws.append(["", "", "", "", f"TOTAL  ({len(api_invoices)} invoices)", "USD", usd_total, f"${usd_total:.2f}"])
    tr = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=tr, column=col)
        cell.font      = total_font
        cell.fill      = total_fill
        cell.border    = border
        cell.alignment = center if col in (1, 6, 7) else left
        if col == 7:
            cell.number_format = currency_fmt

    if cad_total > 0:
        ws.append(["", "", "", "", "", "CAD", cad_total, f"CA${cad_total:.2f}"])
        tr2 = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=tr2, column=col)
            cell.font      = total_font
            cell.fill      = total_fill
            cell.border    = border
            cell.alignment = center if col in (1, 6, 7) else left
            if col == 7:
                cell.number_format = currency_fmt

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [5, 22, 18, 18, 55, 10, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    wb.save(out_path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Fetch Anthropic API invoices from Gmail")
    parser.add_argument("--excel", metavar="FILE", help="Also export results to an Excel file")
    args = parser.parse_args()

    print("\nRAAF — Anthropic API Invoice Fetcher")
    print("=" * 60)

    creds   = get_credentials()
    service = build("gmail", "v1", credentials=creds)

    since = (datetime.now(timezone.utc) - timedelta(days=LOOKBACK_DAYS)).strftime("%Y/%m/%d")
    query = f'from:(anthropic) (subject:invoice OR subject:receipt OR subject:billing) after:{since}'

    print(f"\nSearching: {query}\n")

    response = service.users().messages().list(
        userId="me", q=query, maxResults=50
    ).execute()
    messages = response.get("messages", [])

    if not messages:
        print("No messages matched the search query.")
        return

    api_invoices  = []
    skipped_subs  = []

    for ref in messages:
        msg     = service.users().messages().get(userId="me", id=ref["id"], format="full").execute()
        headers = {h["name"]: h["value"] for h in msg["payload"]["headers"]}
        subject = headers.get("Subject", "(no subject)")
        date    = headers.get("Date", "—")
        sender  = headers.get("From", "—")
        body    = _decode_body(msg["payload"])

        if _is_subscription(subject, body):
            skipped_subs.append(subject)
            continue

        api_invoices.append({
            "date":       date,
            "subject":    subject,
            "sender":     sender,
            "amount":     _find_amount(body),
            "invoice_id": _find_invoice_id(subject, body),
            "period":     _find_period(body),
            "preview":    _clean_preview(body),
        })

    # ── Print results ─────────────────────────────────────────────────────────

    print(f"Results: {len(api_invoices)} API invoice(s) found  |  "
          f"{len(skipped_subs)} subscription invoice(s) skipped\n")

    if skipped_subs:
        print("Skipped (Claude.ai subscription):")
        for s in skipped_subs:
            print(f"  • {s}")
        print()

    if not api_invoices:
        print("No Anthropic API invoices found in the last 90 days.")
        return

    print("=" * 60)
    print("  ANTHROPIC API INVOICES — LAST 90 DAYS")
    print("=" * 60)

    for i, inv in enumerate(api_invoices, 1):
        print(f"\n[{i}]  {inv['subject']}")
        print(f"     Date       : {inv['date']}")
        print(f"     Amount     : {inv['amount']}")
        print(f"     Period     : {inv['period']}")
        print(f"     Invoice #  : {inv['invoice_id']}")
        print(f"     From       : {inv['sender']}")
        print(f"     ── Preview ──────────────────────────────────")
        for line in inv["preview"][:12]:
            print(f"     {line}")

    print(f"\n{'='*60}")
    print(f"Total API invoices: {len(api_invoices)}")
    print(f"Token stored at   : config/.gmail_token.json  (not committed)")

    # ── Excel export ──────────────────────────────────────────────────────────
    excel_path = Path(args.excel) if args.excel else \
        PROJECT_ROOT / f"anthropic_invoices_{datetime.now().strftime('%Y%m%d')}.xlsx"
    export_excel(api_invoices, excel_path)
    print(f"\nExcel saved → {excel_path}")


if __name__ == "__main__":
    main()
